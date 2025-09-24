


from openpyxl.styles import Font, Border, Alignment # , PatternFill
# from openpyxl.formatting.rule import FormulaRule
# from openpyxl.styles import NamedStyle






def format_sheet(sheet):
    # fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    # fill_missing = PatternFill(start_color='eedd00',end_color='eedd00',fill_type='solid')
    # color_shaded =  Font(color='444444')
    # substitutes = {
    #     **sheet_variables.column_letters
    # }
    # sheet.conditional_formatting.add('${col_shortname}$2:${col_shortname}$999999'.format(**substitutes),FormulaRule(formula=['=AND(NOT(ISBLANK(${col_question}2)),NOT(ISBLANK(${col_shortname}2)),NOT(ISBLANK(${col_include}2)),(${col_validation}2="Failed"))'.format(**substitutes)],fill=fill_failed))
    # sheet.conditional_formatting.add('${col_shortname}$2:${col_shortname}$999999'.format(**substitutes),FormulaRule(formula=['=AND(NOT(ISBLANK(${col_question}2)),ISBLANK(${col_shortname}2),NOT(ISBLANK(${col_include}2)))'.format(**substitutes)],fill=fill_missing))
    # sheet.column_dimensions['A'].width = 45
    # sheet.column_dimensions['B'].width = 30
    # sheet.column_dimensions['C'].width = 30
    # sheet.column_dimensions['D'].width = 30
    # sheet.column_dimensions['E'].width = 50
    # sheet.column_dimensions['F'].width = 20
    # sheet.column_dimensions['G'].width = 20
    # sheet.column_dimensions['H'].width = 50
    # for row_num_within_data_range_zero_based, row in enumerate([r for r in sheet.rows][1:]):
    #     # row_num_openpyxl = row_num_within_data_range_zero_based + 1
    #     for cell_index, cell in enumerate(row):
    #         if cell_index==1:
    #             cell.font = color_shaded
    #         elif cell_index==2:
    #             cell.font = color_shaded

    # normal_style = NamedStyle(name='Normal')
    # wb = sheet.parent
    # normal_style = wb.named_styles["Normal"]

    # alignment_left_top = Alignment(indent=2,horizontal='left',vertical='top')
    alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
    alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')

    # Reset styles for the leftmost column (e.g., column A)
    for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
        for cell in row:
            # cell.style = normal_style
            cell.font = Font(bold=False)
            cell.border = Border()
            cell.alignment = alignment_left_top
    
    # Reset styles for the topmost row (e.g., row 1)
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            # cell.style = normal_style
            cell.font = Font(bold=True)
            cell.border = Border()
            cell.alignment = alignment_left_bottom

    for col in sheet.columns:
        col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
        sheet.column_dimensions[col_letter].width = 80
    sheet.column_dimensions['A'].width = 35
    sheet.row_dimensions[1].height = 21




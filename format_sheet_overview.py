
# import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule



def format_sheet(sheet):
    # fill_main = PatternFill(start_color='015254',end_color='015254',fill_type='solid')
    # fill_header = PatternFill(start_color='013E40',end_color='013E40',fill_type='solid')
    # fill_failed = PatternFill(start_color='ee0000',end_color='ee0000',fill_type='solid')
    # font_main = Font(color='ffffff')
    font_main_transparent = Font(color='FFFFFF')
    fill_main_transparent = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    # font_header = Font(color='013E40')
    # border_main = Border(left=Side(style='thick',color='ffffff'),right=Side(style='thick',color='ffffff'),bottom=Side(style='thick',color='ffffff'),top=Side(style='thick',color='ffffff'),)
    border_noborder = Border()
    # # alignment_center = Alignment(horizontal='center',vertical='top')
    # # alignment_indent = Alignment(indent=2,vertical='top')
    alignment_col1 = Alignment(indent=2,horizontal='right',vertical='top')
    alignment_col2 = Alignment(indent=2,horizontal='left',vertical='top')

    font_header = Font(color="666666",bold=True,size=26)

    # df.style.apply(style_header,subset=pd.IndexSlice[:1,])
    # sheet.column_dimensions[[s for s in sheet.columns][0].column_letter].width = 200
    # sheet.column_dimensions[[s for s in sheet.columns][1].column_letter].width = 400    
    # for row in [r for r in sheet.rows][0:1]:
    #     for cell in row:
    #         cell.fill = fill_header
    #         cell.font = font_header
    #         row[0].alignment = alignment_indent
    #         row[1].alignment = alignment_center

    for row_number, row in enumerate([r for r in sheet.rows][1:]):
        sheet.row_dimensions[row_number+2].height = 25
        for cell in row:
            # cell.fill = fill_main
            # cell.font = font_main
            cell.border = border_noborder
            row[0].alignment = alignment_col1
            row[1].alignment = alignment_col2
    for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=26):
        for cell in row:
            cell.fill = fill_main_transparent # fill_header
            cell.font = font_main_transparent # font_header
            cell.border = border_noborder
    # for row in sheet.iter_rows(min_row=2,max_row=999,min_col=1,max_col=26):
    #     for cell in row:
    #         # cell.fill = fill_main
    #         # cell.font = font_main
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 125    
    sheet.row_dimensions[1].height = 65
    sheet.row_dimensions[2].height = 45
    sheet['A1'].fill = fill_main_transparent
    sheet['B1'].fill = fill_main_transparent
    sheet['A1'].font = font_main_transparent
    sheet['B1'].font = font_main_transparent
    sheet['A2'].border = border_noborder
    sheet['B2'].border = border_noborder

    sheet["B2"].font = font_header

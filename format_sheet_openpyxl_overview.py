from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
def format_sheet(sheet):
    font_main_transparent = Font(color='FFFFFF')
    fill_main_transparent = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border_noborder = Border()
    alignment_col1 = Alignment(indent=2,horizontal='right',vertical='top')
    alignment_col2 = Alignment(indent=2,horizontal='left',vertical='top')
    font_header = Font(color="666666",bold=True,size=24)
    for row_number, row in enumerate([r for r in sheet.rows][1:]):
        sheet.row_dimensions[row_number+2].height = 25
        for cell in row:
            cell.border = border_noborder
            row[0].alignment = alignment_col1
            row[1].alignment = alignment_col2
    for row in sheet.iter_rows(min_row=1,max_row=1,min_col=1,max_col=26):
        for cell in row:
            cell.fill = fill_main_transparent # fill_header
            cell.font = font_main_transparent # font_header
            cell.border = border_noborder
    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 125    
    sheet.row_dimensions[1].height = 65
    sheet.row_dimensions[2].height = 45
    sheet['A1'].fill = fill_main_transparent
    sheet['B1'].fill = fill_main_transparent
    sheet['A1'].font = font_main_transparent
    sheet['B1'].font = font_main_transparent
    sheet['A2'].border = border_noborder
    sheet['B2'].border = border_noborder
    sheet["B2"].font = font_header

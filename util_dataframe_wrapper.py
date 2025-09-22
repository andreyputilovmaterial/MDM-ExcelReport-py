
import pandas as pd


from openpyxl.utils import get_column_letter as openpyxl_get_column_letter


# AP 9/21/2025 copied from AA-revival


# this is a helper class
# a wrapper around pandas dataframe
# to make appending or inserting new rows faster
# when we add or insert a row to a pandas df object, it's slow - pandas is creating a copy every time, and is checking for indexes
# instead, we'll be building our data sheet in simple python objects - lists and dicts
# which is fast
# and we'll convert it to a pandas df in one op, which is fast


class PandasDataframeWrapper:
    def __init__(self,columns):
        self._columns = columns
        self._df = pd.DataFrame(
            data =  { col[0]: col[1] for col in zip(self._columns,[[]]*len(self._columns)) },
            index = None,
        )
        self._data = []
        # self._row = 2

    def append(self,*seq):
        row_add = {col[0]: col[1] for col in zip(self._columns,seq)}
        self._data.append(row_add)
    
    def get_working_row_number(self):
        return len(self._data) + 2

    def to_df(self):
        self._df = pd.concat([self._df,pd.DataFrame(self._data)],ignore_index=True)
        index = self.get_index()
        if index:
            self._df.set_index(index,inplace=True)
        return self._df
    
    def get_index(self):
        return self._columns[0] if len(self._columns)>0 else None

    def get_column_letter(self,name):
        index = self._columns.index(name)
        return openpyxl_get_column_letter(index+1)


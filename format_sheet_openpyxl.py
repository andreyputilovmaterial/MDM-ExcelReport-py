from openpyxl.styles import Font, Border, Alignment # , PatternFill
def format_sheet(sheet):
    alignment_left_top = Alignment(indent=0,horizontal='left',vertical='top')
    alignment_left_bottom = Alignment(indent=0,horizontal='left',vertical='bottom')
    for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.font = Font(bold=False)
            cell.border = Border()
            cell.alignment = alignment_left_top
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        for cell in col:
            cell.font = Font(bold=True)
            cell.border = Border()
            cell.alignment = alignment_left_bottom
    for col in sheet.columns:
        col_letter = col[0].column_letter  # Get the column letter (e.g., 'A', 'B', etc.)
        sheet.column_dimensions[col_letter].width = 80
    sheet.column_dimensions['A'].width = 35
    sheet.row_dimensions[1].height = 21
